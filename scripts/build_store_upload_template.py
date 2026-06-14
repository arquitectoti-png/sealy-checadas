from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation


OUTPUT = Path("outputs/godaddy_prueba/admin/assets/layout_carga_tiendas.xlsx")


def style_header(cell):
    cell.font = Font(color="FFFFFF", bold=True)
    cell.fill = PatternFill("solid", fgColor="1F5FBF")
    cell.alignment = Alignment(horizontal="center")


def main():
    wb = Workbook()
    ws = wb.active
    ws.title = "Tiendas"

    headers = ["cadena", "nombre", "direccion", "latitud", "longitud", "radio", "zona_horaria"]
    examples = [
        ["Coppel", "Hermosillo Centro", "Direccion completa de la tienda", 29.072967, -110.955919, 50, "AUTO"],
        ["Suburbia", "Monterrey Centro", "Direccion completa de la tienda", 25.686614, -100.316113, 50, "AUTO"],
        ["Coppel", "Tijuana Centro", "Direccion completa de la tienda", 32.514947, -117.038247, 50, "AUTO"],
    ]
    comments = {
        "cadena": "Ejemplo: Coppel, Suburbia, Liverpool. Obligatorio.",
        "nombre": "Nombre unico de la tienda. Si ya existe, se actualiza. Obligatorio.",
        "direccion": "Direccion completa. Obligatorio.",
        "latitud": "Coordenada decimal. Ejemplo: 29.072967. Obligatorio.",
        "longitud": "Coordenada decimal. Ejemplo: -110.955919. Obligatorio.",
        "radio": "Radio permitido en metros. Recomendado: 50. Obligatorio.",
        "zona_horaria": "Usa AUTO para calcular por coordenadas. Tambien puedes usar: CENTRO, SONORA, PACIFICO, BAJA o CANCUN.",
    }

    ws.append(headers)
    for row in examples:
        ws.append(row)

    thin = Side(style="thin", color="D7DEE8")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    widths = [18, 28, 44, 14, 14, 10, 18]
    max_rows = 503

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:G{max_rows}"
    ws.sheet_view.showGridLines = False

    for col, header in enumerate(headers, 1):
        cell = ws.cell(1, col)
        style_header(cell)
        cell.border = border
        cell.comment = Comment(comments[header], "Sealy")
        ws.column_dimensions[get_column_letter(col)].width = widths[col - 1]

    for row in ws.iter_rows(min_row=2, max_row=max_rows, min_col=1, max_col=7):
        for cell in row:
            cell.border = border

    for row in range(2, max_rows + 1):
        ws.cell(row, 4).number_format = "0.000000"
        ws.cell(row, 5).number_format = "0.000000"
        ws.cell(row, 6).number_format = "0"

    validation = DataValidation(
        type="list",
        formula1='"AUTO,CENTRO,SONORA,PACIFICO,BAJA,CANCUN"',
        allow_blank=True,
    )
    validation.promptTitle = "Zona horaria"
    validation.prompt = "Selecciona AUTO o una zona manual: CENTRO, SONORA, PACIFICO, BAJA o CANCUN"
    validation.errorTitle = "Zona horaria invalida"
    validation.error = "Usa una clave valida del catalogo de zonas horarias."
    ws.add_data_validation(validation)
    validation.add(f"G2:G{max_rows}")

    catalog = wb.create_sheet("Catalogo zonas")
    catalog_headers = ["clave", "descripcion", "zona_tecnica", "usar_para"]
    catalog_rows = [
        ["AUTO", "Automatico por coordenadas", "Calculado por el sistema", "Recomendado para la carga masiva normal"],
        ["CENTRO", "Centro / Monterrey / CDMX", "America/Mexico_City", "CDMX, Monterrey, Guadalajara, Puebla, Queretaro, Merida"],
        ["SONORA", "Sonora / Hermosillo", "America/Hermosillo", "Hermosillo y estado de Sonora"],
        ["PACIFICO", "Pacifico / Sinaloa / BCS", "America/Mazatlan", "Sinaloa, Culiacan, Mazatlan, La Paz, BCS"],
        ["BAJA", "Baja California", "America/Tijuana", "Tijuana, Mexicali, Baja California"],
        ["CANCUN", "Quintana Roo / Cancun", "America/Cancun", "Cancun y Quintana Roo"],
    ]
    catalog.append(catalog_headers)
    for row in catalog_rows:
        catalog.append(row)

    catalog.freeze_panes = "A2"
    catalog.auto_filter.ref = "A1:D7"
    catalog.sheet_view.showGridLines = False
    catalog_widths = [16, 32, 28, 60]
    for col, header in enumerate(catalog_headers, 1):
        cell = catalog.cell(1, col)
        style_header(cell)
        cell.border = border
        catalog.column_dimensions[get_column_letter(col)].width = catalog_widths[col - 1]

    for row in catalog.iter_rows(min_row=2, max_row=7, min_col=1, max_col=4):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUTPUT)

    check = load_workbook(OUTPUT)
    assert check.sheetnames == ["Tiendas", "Catalogo zonas"]
    assert check["Tiendas"]["A1"].value == "cadena"
    print(OUTPUT.resolve())


if __name__ == "__main__":
    main()
