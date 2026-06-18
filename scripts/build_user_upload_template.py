from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


OUTPUT = Path("outputs/godaddy_prueba/admin/assets/layout_carga_usuarios.xlsx")


def style_header(cell):
    cell.font = Font(color="FFFFFF", bold=True)
    cell.fill = PatternFill("solid", fgColor="1F5FBF")
    cell.alignment = Alignment(horizontal="center")


def main():
    wb = Workbook()
    ws = wb.active
    ws.title = "Promotores"

    headers = ["nombre", "email", "numero_empleado", "rfc", "telefono", "supervisor", "contrasena"]
    examples = [
        ["Promotor 101", "promotor101@sealy.com", "PRO101", "PRO101010AAA", "5512340001", "", "Cambiar123!"],
        ["Promotor 102", "promotor102@sealy.com", "PRO102", "PRO102010AAA", "5512340002", "Supervisor 1", "Cambiar123!"],
        ["Promotor 103", "promotor103@sealy.com", "PRO103", "PRO103010AAA", "5512340003", "supervisor1@sealy.com", "Cambiar123!"],
    ]
    comments = {
        "nombre": "Nombre completo del promotor. Obligatorio.",
        "email": "Correo electronico del promotor. Debe ser unico y valido. Obligatorio.",
        "numero_empleado": "Clave o numero de empleado. Si ya existe, el usuario se actualiza. Obligatorio.",
        "rfc": "RFC del promotor. Obligatorio para historico y busqueda.",
        "telefono": "Telefono del promotor. Obligatorio.",
        "supervisor": "Opcional. Puede ser nombre, email o numero de empleado del supervisor. Los supervisores pueden ver a todos los promotores.",
        "contrasena": "Opcional si se define contrasena default en el panel. Minimo recomendado: 8 caracteres.",
    }

    ws.append(headers)
    for row in examples:
        ws.append(row)

    thin = Side(style="thin", color="D7DEE8")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    widths = [30, 32, 20, 18, 18, 28, 18]
    max_rows = 503

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:G{max_rows}"
    ws.sheet_view.showGridLines = False

    for col, header in enumerate(headers, 1):
        cell = ws.cell(1, col)
        style_header(cell)
        cell.border = border
        cell.comment = Comment(comments[header], "Sealy")
        ws.column_dimensions[get_column_letter(col)].width = widths[col - 1]

    for row in ws.iter_rows(min_row=2, max_row=max_rows, min_col=1, max_col=len(headers)):
        for cell in row:
            cell.border = border

    catalog = wb.create_sheet("Instrucciones")
    instruction_rows = [
        ["Campo", "Obligatorio", "Descripcion"],
        ["nombre", "Si", "Nombre completo del promotor."],
        ["email", "Si", "Correo electronico valido."],
        ["numero_empleado", "Si", "Clave unica. Si ya existe, actualiza al promotor."],
        ["rfc", "Si", "RFC obligatorio para identificar historico aunque se repita el nombre."],
        ["telefono", "Si", "Telefono del promotor."],
        ["supervisor", "No", "Referencia opcional. Puede ser nombre, email o numero de empleado del supervisor."],
        ["contrasena", "No", "Si se deja vacia, se usa la contrasena default capturada en el panel."],
    ]
    for row in instruction_rows:
        catalog.append(row)

    catalog.freeze_panes = "A2"
    catalog.auto_filter.ref = "A1:C8"
    catalog.sheet_view.showGridLines = False
    catalog_widths = [22, 14, 80]
    for col in range(1, 4):
        cell = catalog.cell(1, col)
        style_header(cell)
        catalog.column_dimensions[get_column_letter(col)].width = catalog_widths[col - 1]

    for row in catalog.iter_rows(min_row=1, max_row=8, min_col=1, max_col=3):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUTPUT)

    check = load_workbook(OUTPUT)
    assert check.sheetnames == ["Promotores", "Instrucciones"]
    assert check["Promotores"]["A1"].value == "nombre"
    print(OUTPUT.resolve())


if __name__ == "__main__":
    main()
